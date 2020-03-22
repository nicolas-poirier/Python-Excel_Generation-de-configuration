#! python3
# coding=utf-8
"""
Contexte: GENERAL
Génération de fichiers de configurations avec utilisation de template
"""
import openpyxl
import os
import shutil
import re
import ipaddress

if not os.path.exists("output"):
    os.makedirs("output")
else:
    shutil.rmtree("output")
    os.makedirs("output")

# le fichier Excel avec les infos
wb = openpyxl.load_workbook("input.xlsx")

# onglet du fichier Excel contenant la conf des switchs
# possibilité d'ajouter d'autres onglets sur le même modèle
onglet_info_sw = wb["CONF-SWITCH"]

nbre_sw = onglet_info_sw.max_row + 1
for i in range(2, nbre_sw):
    dico_variables = {
        "variables": {}
    }
    for j in range(1, onglet_info_sw.max_column + 1):        
        isIP = True
        try:
            ipaddress.ip_address(str(onglet_info_sw.cell(row=i, column=j).value))
            isIp = True
        except:
            isIP = False        
        if str(onglet_info_sw.cell(row=1, column=j).value) == "TEMPLATE":
            dico_variables["variables"][onglet_info_sw.cell(row=1, column=j).value] = \
                str(onglet_info_sw.cell(row=i, column=j).value)
        elif isIP:
            dico_variables["variables"][onglet_info_sw.cell(row=1, column=j).value] = \
                str(onglet_info_sw.cell(row=i, column=j).value)
        else:
            dico_variables["variables"][onglet_info_sw.cell(row=1, column=j).value] = \
                str(onglet_info_sw.cell(row=i, column=j).value).replace(".",",")

    # utilisation du template spécifié
    with open(dico_variables["variables"]["TEMPLATE"]) as f:
        s = f.read()
    with open("output\\" + dico_variables["variables"]["HOSTNAME"] + ".txt", "w") as f:
        # remplacement de chaque variable dans le fichier cible
        for var in dico_variables["variables"]:
            s = re.sub("\$" + var, dico_variables["variables"][var], s)
        f.write(s)
